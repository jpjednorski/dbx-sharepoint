from __future__ import annotations

import io
import json
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
import responses
from dbx_sharepoint.client import SharePointClient
from dbx_sharepoint.exceptions import (
    SharePointAuthError,
    SharePointFileNotFoundError,
    SharePointPermissionError,
    SharePointThrottledError,
)


@pytest.fixture
def mock_credential():
    cred = MagicMock()
    token = MagicMock()
    token.token = "fake-access-token"
    cred.get_token.return_value = token
    return cred


@pytest.fixture
def gov_client(mock_credential):
    return SharePointClient(
        credential=mock_credential,
        site_url="https://myorg.sharepoint.us/sites/TeamSite",
    )


@pytest.fixture
def commercial_client(mock_credential):
    return SharePointClient(
        credential=mock_credential,
        site_url="https://myorg.sharepoint.com/sites/TeamSite",
    )


class TestSharePointClientInit:
    def test_gov_auto_detection(self, gov_client):
        assert gov_client._graph_endpoint == "https://graph.microsoft.us"

    def test_commercial_auto_detection(self, commercial_client):
        assert commercial_client._graph_endpoint == "https://graph.microsoft.com"

    def test_explicit_graph_endpoint_overrides(self, mock_credential):
        client = SharePointClient(
            credential=mock_credential,
            site_url="https://myorg.sharepoint.us/sites/TeamSite",
            graph_endpoint="https://custom.graph.endpoint",
        )
        assert client._graph_endpoint == "https://custom.graph.endpoint"

    def test_site_name_parsed(self, gov_client):
        assert gov_client._site_name == "TeamSite"

    def test_hostname_parsed(self, gov_client):
        assert gov_client._hostname == "myorg.sharepoint.us"


class TestFromDatabricksSecrets:
    def test_factory_method(self):
        mock_dbutils = MagicMock()
        mock_dbutils.secrets.get.side_effect = lambda scope, key: {
            "prod-tenant-id": "t123",
            "prod-client-id": "c456",
            "prod-client-secret": "s789",
            "prod-site-url": "https://myorg.sharepoint.us/sites/Team",
        }[key]

        with patch("dbx_sharepoint.auth.ClientSecretCredential") as mock_cred:
            mock_cred.return_value = MagicMock()
            client = SharePointClient.from_databricks_secrets(
                dbutils=mock_dbutils,
                scope="sharepoint",
                prefix="prod",
            )
            assert client._hostname == "myorg.sharepoint.us"
            # Verify Gov authority was passed to credential
            mock_cred.assert_called_once_with(
                tenant_id="t123",
                client_id="c456",
                client_secret="s789",
                authority="https://login.microsoftonline.us",
            )

    def test_factory_site_url_param_overrides_secret(self):
        mock_dbutils = MagicMock()
        secrets = {
            "tenant-id": "t123",
            "client-id": "c456",
            "client-secret": "s789",
            "site-url": "https://secret.sharepoint.us/sites/FromSecret",
        }

        def get_secret(scope, key):
            if key not in secrets:
                raise Exception(f"Secret {key} not found")
            return secrets[key]

        mock_dbutils.secrets.get.side_effect = get_secret

        with patch("dbx_sharepoint.auth.ClientSecretCredential") as mock_cred:
            mock_cred.return_value = MagicMock()
            client = SharePointClient.from_databricks_secrets(
                dbutils=mock_dbutils,
                scope="sharepoint",
                site_url="https://param.sharepoint.us/sites/FromParam",
            )
            assert client._site_name == "FromParam"


class TestListFiles:
    @responses.activate
    def test_list_files_from_url(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/reports:/children",
            json={
                "value": [
                    {
                        "name": "q1.xlsx",
                        "webUrl": "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/reports/q1.xlsx",
                        "size": 12345,
                        "lastModifiedDateTime": "2026-01-15T10:30:00Z",
                        "lastModifiedBy": {"user": {"displayName": "Jane Doe"}},
                        "file": {},
                    },
                    {
                        "name": "subfolder",
                        "webUrl": "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/reports/subfolder",
                        "size": 0,
                        "lastModifiedDateTime": "2026-02-01T08:00:00Z",
                        "lastModifiedBy": {"user": {"displayName": "John Smith"}},
                        "folder": {"childCount": 3},
                    },
                ]
            },
        )

        df = gov_client.list_files(
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/reports/"
        )

        assert isinstance(df, pd.DataFrame)
        assert list(df.columns) == ["name", "path", "size_bytes", "modified_at", "modified_by", "is_folder"]
        assert len(df) == 2
        assert df.iloc[0]["name"] == "q1.xlsx"
        assert df.iloc[1]["is_folder"] == True

    @responses.activate
    def test_list_files_follows_pagination(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        next_link = "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/big:/children?$skiptoken=XYZ"
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/big:/children",
            json={
                "value": [{"name": f"f{i}.txt", "size": i, "file": {}} for i in range(2)],
                "@odata.nextLink": next_link,
            },
        )
        responses.add(
            responses.GET,
            next_link,
            json={"value": [{"name": f"f{i}.txt", "size": i, "file": {}} for i in range(2, 5)]},
            match_querystring=True,
        )

        df = gov_client.list_files("/big")
        assert len(df) == 5
        assert list(df["name"]) == [f"f{i}.txt" for i in range(5)]


class TestDownload:
    @responses.activate
    def test_download_file(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/file.pdf:/content",
            body=b"fake-pdf-content",
            content_type="application/pdf",
        )

        content = gov_client.download(
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/file.pdf"
        )
        assert content == b"fake-pdf-content"


class TestUpload:
    @responses.activate
    def test_upload_file(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.PUT,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/output.xlsx:/content",
            json={"id": "item-id-456", "name": "output.xlsx"},
        )

        gov_client.upload(
            b"fake-excel-bytes",
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/output.xlsx",
        )

        assert responses.calls[1].request.body == b"fake-excel-bytes"

    @responses.activate
    def test_upload_large_file_uses_upload_session(self, gov_client):
        # Force tiny threshold/chunk so a 10-byte payload exercises 3 chunks
        gov_client._UPLOAD_SESSION_THRESHOLD = 3
        gov_client._UPLOAD_CHUNK_SIZE = 4

        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.POST,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/big.xlsx:/createUploadSession",
            json={"uploadUrl": "https://upload.sharepoint.us/session/abc"},
        )
        responses.add(
            responses.PUT,
            "https://upload.sharepoint.us/session/abc",
            json={"nextExpectedRanges": ["4-9"]},
            status=202,
        )
        responses.add(
            responses.PUT,
            "https://upload.sharepoint.us/session/abc",
            json={"nextExpectedRanges": ["8-9"]},
            status=202,
        )
        responses.add(
            responses.PUT,
            "https://upload.sharepoint.us/session/abc",
            json={"id": "item-final"},
            status=201,
        )

        payload = b"0123456789"
        gov_client.upload(payload, "/Shared Documents/big.xlsx")

        chunk_calls = [c for c in responses.calls if c.request.url.startswith("https://upload.")]
        assert len(chunk_calls) == 3
        assert chunk_calls[0].request.headers["Content-Range"] == "bytes 0-3/10"
        assert chunk_calls[1].request.headers["Content-Range"] == "bytes 4-7/10"
        assert chunk_calls[2].request.headers["Content-Range"] == "bytes 8-9/10"
        # Upload URL is pre-signed — client must not send Authorization on chunk PUTs
        assert "Authorization" not in chunk_calls[0].request.headers


class TestErrorHandling:
    @responses.activate
    def test_401_raises_auth_error(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"error": {"message": "Unauthorized"}},
            status=401,
        )

        with pytest.raises(SharePointAuthError):
            gov_client.list_files(
                "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/"
            )

    @responses.activate
    def test_404_raises_file_not_found(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/missing.xlsx:/content",
            json={"error": {"message": "Not found"}},
            status=404,
        )

        with pytest.raises(SharePointFileNotFoundError):
            gov_client.download(
                "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/missing.xlsx"
            )

    @responses.activate
    def test_403_raises_permission_error(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"error": {"message": "Forbidden"}},
            status=403,
        )

        with pytest.raises(SharePointPermissionError):
            gov_client.list_files(
                "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/"
            )

    @responses.activate
    def test_429_retries_then_succeeds(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        # First call: throttled
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/file.pdf:/content",
            json={"error": {"message": "Too many requests"}},
            status=429,
            headers={"Retry-After": "1"},
        )
        # Second call: succeeds
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/file.pdf:/content",
            body=b"content",
            content_type="application/pdf",
        )

        content = gov_client.download(
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/file.pdf"
        )
        assert content == b"content"


class TestReadExcel:
    @responses.activate
    def test_read_excel_returns_dataframe(self, gov_client):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "name"
        ws["B1"] = "value"
        ws["A2"] = "alpha"
        ws["B2"] = 100
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/data.xlsx:/content",
            body=xlsx_bytes,
            content_type="application/octet-stream",
        )

        df = gov_client.read_excel(
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/data.xlsx"
        )
        assert list(df.columns) == ["name", "value"]
        assert df.iloc[0]["name"] == "alpha"


class TestWriteExcel:
    @responses.activate
    def test_write_excel_uploads(self, gov_client):
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.PUT,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/out.xlsx:/content",
            json={"id": "item-id-789"},
        )

        df = pd.DataFrame({"col": [1, 2, 3]})
        gov_client.write_excel(
            df, "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/out.xlsx"
        )

        uploaded = responses.calls[1].request.body
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(uploaded))
        assert wb.active["A1"].value == "col"


class TestOpenTemplateAndSave:
    @responses.activate
    def test_open_template_and_save(self, gov_client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        ws["A1"] = "Title Placeholder"
        buf = io.BytesIO()
        wb.save(buf)
        template_bytes = buf.getvalue()

        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/myorg.sharepoint.us:/sites/TeamSite",
            json={"id": "site-id-123"},
        )
        responses.add(
            responses.GET,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/template.xlsx:/content",
            body=template_bytes,
            content_type="application/octet-stream",
        )
        responses.add(
            responses.PUT,
            "https://graph.microsoft.us/v1.0/sites/site-id-123/drive/root:/Shared Documents/filled.xlsx:/content",
            json={"id": "item-id-filled"},
        )

        template = gov_client.open_template(
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/template.xlsx"
        )
        template.set_value("Report", cell="A1", value="Q1 2026")

        gov_client.save(
            template,
            "https://myorg.sharepoint.us/sites/TeamSite/Shared Documents/filled.xlsx",
        )

        uploaded = responses.calls[2].request.body
        wb2 = openpyxl.load_workbook(io.BytesIO(uploaded))
        assert wb2["Report"]["A1"].value == "Q1 2026"
