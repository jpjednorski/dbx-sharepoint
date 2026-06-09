from __future__ import annotations

import io
import zipfile

import openpyxl
import pandas as pd
import pytest
from dbx_sharepoint.excel import (
    dataframe_from_excel_bytes,
    dataframe_to_excel_bytes,
    dataframe_to_excel_bytes_from_template,
    Template,
)


def _make_xlsm_bytes(sheet_name: str = "Sheet1") -> bytes:
    """Helper: create macro-enabled (.xlsm) workbook bytes with a VBA project.

    openpyxl can't author a vbaProject.bin, so we build a normal workbook and
    splice a stub xl/vbaProject.bin into the zip container. That is enough to
    exercise the keep_vba round-trip: the marker must survive a save.
    """
    wb = openpyxl.Workbook()
    wb.active.title = sheet_name
    buf = io.BytesIO()
    wb.save(buf)

    src = io.BytesIO(buf.getvalue())
    out = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
        for item in zin.namelist():
            zout.writestr(item, zin.read(item))
        zout.writestr("xl/vbaProject.bin", b"\xcf\x11\xe0vba-stub")
    return out.getvalue()


def _make_blank_xlsm_bytes(sheet_name: str = "Sheet1") -> bytes:
    """Helper: macro-enabled (.xlsm) workbook with NO xl/vbaProject.bin.

    Reproduces a blank macro-enabled template as Excel saves it: the package
    carries the macroEnabled content type, but no VBA binary exists until a
    macro is added. Detecting macro-enabled by vbaProject.bin alone misses
    this case and yields a file Excel rejects as corrupt.
    """
    wb = openpyxl.Workbook()
    wb.active.title = sheet_name
    buf = io.BytesIO()
    wb.save(buf)

    src = io.BytesIO(buf.getvalue())
    out = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
        for item in zin.namelist():
            content = zin.read(item)
            if item == "[Content_Types].xml":
                content = content.replace(
                    b"application/vnd.openxmlformats-officedocument."
                    b"spreadsheetml.sheet.main+xml",
                    b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
            zout.writestr(item, content)
    return out.getvalue()


def _workbook_content_type(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        ct = zf.read("[Content_Types].xml").decode()
    import re

    m = re.search(r'workbook\.xml" ContentType="([^"]+)"', ct)
    return m.group(1) if m else ""


def _make_workbook_bytes(data: dict, sheet_name: str = "Sheet1") -> bytes:
    """Helper: create an xlsx in memory from a dict of column->values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(data.keys())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        for row_idx, val in enumerate(data[header], 2):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_template_bytes_with_named_range() -> bytes:
    """Helper: create a template with a named range 'data_table' covering B2:D4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Report Title"
    ws["B1"] = "Col1"
    ws["C1"] = "Col2"
    ws["D1"] = "Col3"
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName("data_table", attr_text="Report!$B$2:$D$4")
    wb.defined_names.add(defn)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestDataframeFromExcelBytes:
    def test_read_default_sheet(self):
        data = {"name": ["Alice", "Bob"], "age": [30, 25]}
        xlsx_bytes = _make_workbook_bytes(data)
        df = dataframe_from_excel_bytes(xlsx_bytes)
        assert list(df.columns) == ["name", "age"]
        assert len(df) == 2
        assert df.iloc[0]["name"] == "Alice"

    def test_read_specific_sheet(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1["A1"] = "x"
        ws1["A2"] = 1
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "y"
        ws2["A2"] = 2
        buf = io.BytesIO()
        wb.save(buf)

        df = dataframe_from_excel_bytes(buf.getvalue(), sheet_name="Second")
        assert list(df.columns) == ["y"]
        assert df.iloc[0]["y"] == 2


class TestDataframeToExcelBytes:
    def test_roundtrip(self):
        df = pd.DataFrame({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
        xlsx_bytes = dataframe_to_excel_bytes(df, sheet_name="Results")
        df2 = dataframe_from_excel_bytes(xlsx_bytes, sheet_name="Results")
        assert list(df2.columns) == ["col1", "col2"]
        assert len(df2) == 3

    def test_default_sheet_name(self):
        df = pd.DataFrame({"a": [1]})
        xlsx_bytes = dataframe_to_excel_bytes(df)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert "Sheet1" in wb.sheetnames


class TestTemplateFillRange:
    def test_fill_with_start_cell(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Data")
        template = Template(template_bytes)

        df = pd.DataFrame({"x": [10, 20], "y": [30, 40]})
        template.fill_range("Data", start_cell="A2", data=df)

        wb = template._workbook
        ws = wb["Data"]
        assert ws["A2"].value == 10
        assert ws["B2"].value == 30
        assert ws["A3"].value == 20
        assert ws["B3"].value == 40

    def test_fill_with_end_cell_data_fits(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Data")
        template = Template(template_bytes)

        df = pd.DataFrame({"x": [1, 2], "y": [3, 4]})
        template.fill_range("Data", start_cell="A1", end_cell="B2", data=df)

        ws = template._workbook["Data"]
        assert ws["A1"].value == 1
        assert ws["B2"].value == 4

    def test_fill_with_end_cell_data_exceeds_raises(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Data")
        template = Template(template_bytes)

        df = pd.DataFrame({"x": [1, 2, 3], "y": [4, 5, 6]})
        with pytest.raises(ValueError, match="exceeds"):
            template.fill_range("Data", start_cell="A1", end_cell="B2", data=df)

    def test_fill_with_orientation_columns(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Data")
        template = Template(template_bytes)

        df = pd.DataFrame({"x": [10, 20], "y": [30, 40]})
        template.fill_range("Data", start_cell="A1", data=df, orientation="columns")

        ws = template._workbook["Data"]
        # Transposed: each df row becomes a column
        assert ws["A1"].value == 10  # row 0, col x
        assert ws["A2"].value == 30  # row 0, col y
        assert ws["B1"].value == 20  # row 1, col x
        assert ws["B2"].value == 40  # row 1, col y

    def test_fill_named_range(self):
        template_bytes = _make_template_bytes_with_named_range()
        template = Template(template_bytes)

        df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6], "c": [7, 8, 9]})
        template.fill_range(named_range="data_table", data=df)

        ws = template._workbook["Report"]
        assert ws["B2"].value == 1
        assert ws["C2"].value == 4
        assert ws["D4"].value == 9

    def test_fill_named_range_exceeds_without_expand(self):
        template_bytes = _make_template_bytes_with_named_range()
        template = Template(template_bytes)

        df = pd.DataFrame({"a": [1, 2, 3, 4], "b": [5, 6, 7, 8], "c": [9, 10, 11, 12]})
        with pytest.raises(ValueError, match="exceeds"):
            template.fill_range(named_range="data_table", data=df)

    def test_fill_invalid_orientation_raises(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Data")
        template = Template(template_bytes)
        df = pd.DataFrame({"x": [1]})
        with pytest.raises(ValueError, match="orientation"):
            template.fill_range("Data", start_cell="A1", data=df, orientation="row")

    def test_fill_named_range_expand_allowed(self):
        template_bytes = _make_template_bytes_with_named_range()
        template = Template(template_bytes)

        df = pd.DataFrame({"a": [1, 2, 3, 4], "b": [5, 6, 7, 8], "c": [9, 10, 11, 12]})
        template.fill_range(named_range="data_table", data=df, allow_expand=True)

        ws = template._workbook["Report"]
        assert ws["B5"].value == 4
        assert ws["D5"].value == 12


class TestTemplateSetValue:
    def test_set_single_value(self):
        template_bytes = _make_workbook_bytes({"header": []}, sheet_name="Report")
        template = Template(template_bytes)

        template.set_value("Report", cell="A1", value="Q1 2026 Report")

        ws = template._workbook["Report"]
        assert ws["A1"].value == "Q1 2026 Report"


class TestMacroEnabledTemplate:
    def test_template_autodetects_and_preserves_vba(self):
        template = Template(_make_xlsm_bytes("Macros"))
        assert template._keep_vba is True

        template.set_value("Macros", cell="A1", value=42)
        out = template.to_bytes()

        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            assert "xl/vbaProject.bin" in zf.namelist()

    def test_plain_xlsx_does_not_keep_vba(self):
        template = Template(_make_workbook_bytes({"col": [1]}, sheet_name="S1"))
        assert template._keep_vba is False

    def test_dataframe_to_excel_bytes_from_template_preserves_vba(self):
        df = pd.DataFrame({"x": [1, 2], "y": [3, 4]})
        out = dataframe_to_excel_bytes_from_template(
            _make_xlsm_bytes("Sheet1"), df, start_cell="A1"
        )

        with zipfile.ZipFile(io.BytesIO(out)) as zf:
            assert "xl/vbaProject.bin" in zf.namelist()

        wb = openpyxl.load_workbook(io.BytesIO(out), keep_vba=True)
        ws = wb["Sheet1"]
        assert ws["A1"].value == "x"
        assert ws["A2"].value == 1
        assert ws["B3"].value == 4

    def test_from_template_defaults_to_active_sheet(self):
        df = pd.DataFrame({"x": [9]})
        out = dataframe_to_excel_bytes_from_template(_make_xlsm_bytes("Only"), df)
        wb = openpyxl.load_workbook(io.BytesIO(out), keep_vba=True)
        assert wb["Only"]["A1"].value == "x"
        assert wb["Only"]["A2"].value == 9

    def test_blank_macro_template_detected_without_vba_binary(self):
        # A blank .xlsm has the macroEnabled content type but no vbaProject.bin.
        template = Template(_make_blank_xlsm_bytes("Sheet1"))
        assert template._keep_vba is True

    def test_blank_macro_template_output_keeps_macro_content_type(self):
        # Regression: output must declare the macroEnabled content type, or
        # Excel rejects the .xlsm as "file format or extension is not valid".
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        out = dataframe_to_excel_bytes_from_template(
            _make_blank_xlsm_bytes("Sheet1"), df
        )
        assert _workbook_content_type(out) == (
            "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
        )
        wb = openpyxl.load_workbook(io.BytesIO(out), keep_vba=True)
        assert wb["Sheet1"]["A1"].value == "a"
        assert wb["Sheet1"]["A2"].value == 1
        assert wb["Sheet1"]["B3"].value == 4

    def test_blank_macro_template_renames_only_blank_sheet_for_target(self):
        df = pd.DataFrame({"risk_level": ["HIGH"], "value": [10]})
        out = dataframe_to_excel_bytes_from_template(
            _make_blank_xlsm_bytes("Sheet1"), df, sheet_name="CAPS"
        )

        assert _workbook_content_type(out) == (
            "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
        )
        wb = openpyxl.load_workbook(io.BytesIO(out), keep_vba=True)
        assert wb.sheetnames == ["CAPS"]
        assert wb["CAPS"]["A1"].value == "risk_level"
        assert wb["CAPS"]["A2"].value == "HIGH"
        assert wb["CAPS"]["B2"].value == 10

    def test_template_write_can_omit_headers(self):
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        out = dataframe_to_excel_bytes_from_template(
            _make_blank_xlsm_bytes("Sheet1"), df, include_header=False
        )

        wb = openpyxl.load_workbook(io.BytesIO(out), keep_vba=True)
        assert wb["Sheet1"]["A1"].value == 1
        assert wb["Sheet1"]["B2"].value == 4

    def test_missing_sheet_creates_new_sheet_when_template_has_content(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instructions"
        ws["A1"] = "Do not overwrite"
        buf = io.BytesIO()
        wb.save(buf)

        df = pd.DataFrame({"a": [1]})
        out = dataframe_to_excel_bytes_from_template(
            buf.getvalue(), df, sheet_name="CAPS"
        )

        wb2 = openpyxl.load_workbook(io.BytesIO(out))
        assert wb2.sheetnames == ["Instructions", "CAPS"]
        assert wb2["Instructions"]["A1"].value == "Do not overwrite"
        assert wb2["CAPS"]["A1"].value == "a"
        assert wb2["CAPS"]["A2"].value == 1

    def test_plain_xlsx_output_stays_regular_content_type(self):
        df = pd.DataFrame({"a": [1]})
        out = dataframe_to_excel_bytes_from_template(
            _make_workbook_bytes({"h": []}, sheet_name="Sheet1"), df
        )
        assert _workbook_content_type(out) == (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet.main+xml"
        )


class TestTemplateToBytes:
    def test_to_bytes_returns_valid_xlsx(self):
        template_bytes = _make_workbook_bytes({"col": [1]}, sheet_name="S1")
        template = Template(template_bytes)
        template.set_value("S1", cell="A1", value="modified")

        output = template.to_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(output))
        assert wb["S1"]["A1"].value == "modified"
