from __future__ import annotations

import io

import openpyxl
import pytest
import responses
from dbx_sharepoint.exceptions import SharePointError
from dbx_sharepoint.shared_link import read_excel_from_shared_link


def _make_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["A2"] = 42
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestReadExcelFromSharedLink:
    @responses.activate
    def test_url_without_query_params(self):
        xlsx = _make_xlsx_bytes()
        responses.add(
            responses.GET,
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg",
            body=xlsx,
            content_type="application/octet-stream",
            match_querystring=False,
        )

        df = read_excel_from_shared_link(
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg"
        )
        assert df.iloc[0]["x"] == 42
        assert "download=1" in responses.calls[0].request.url

    @responses.activate
    def test_url_with_existing_query_params(self):
        xlsx = _make_xlsx_bytes()
        responses.add(
            responses.GET,
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg",
            body=xlsx,
            content_type="application/octet-stream",
            match_querystring=False,
        )

        df = read_excel_from_shared_link(
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg?e=abc123"
        )
        assert "&download=1" in responses.calls[0].request.url

    @responses.activate
    def test_specific_sheet(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1["A1"] = "a"
        ws1["A2"] = 1
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "b"
        ws2["A2"] = 99
        buf = io.BytesIO()
        wb.save(buf)

        responses.add(
            responses.GET,
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg",
            body=buf.getvalue(),
            content_type="application/octet-stream",
            match_querystring=False,
        )

        df = read_excel_from_shared_link(
            "https://myorg.sharepoint.us/:x:/s/Team/EaBcDeFg",
            sheet_name="Second",
        )
        assert df.iloc[0]["b"] == 99

    @responses.activate
    def test_html_response_raises_clear_error(self):
        responses.add(
            responses.GET,
            "https://myorg.sharepoint.us/:x:/s/Team/Expired",
            body=b"<html><body>Sign in</body></html>",
            content_type="text/html",
            match_querystring=False,
        )

        with pytest.raises(SharePointError, match="did not return an Excel file"):
            read_excel_from_shared_link(
                "https://myorg.sharepoint.us/:x:/s/Team/Expired"
            )
