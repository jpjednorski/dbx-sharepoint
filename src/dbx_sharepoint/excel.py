from __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from typing import Optional, Union

import openpyxl
import pandas as pd
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

_VALID_ORIENTATIONS = ("rows", "columns")

# Microsoft Information Protection (MIP) sensitivity-label metadata. Modern
# Office (2020+ label format) stores the label in a dedicated package part,
# ``docMetadata/LabelInfo.xml``, wired in through ``[Content_Types].xml`` and
# the root relationships. This is the classification-only form: it marks the
# file with a label but does not encrypt it. openpyxl does not understand this
# part and silently drops it on save, so a labeled template round-tripped
# through openpyxl comes out unlabeled. These helpers re-inject it.
_LABEL_PART = "docMetadata/LabelInfo.xml"
_LABEL_CONTENT_TYPE = "application/vnd.ms-office.classificationlabels+xml"
_LABEL_REL_TYPE = (
    "http://schemas.microsoft.com/office/2020/02/relationships/classificationlabels"
)
_LABEL_NS = "http://schemas.microsoft.com/office/2020/mipLabelMetadata"
_CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class SensitivityLabel:
    """A Microsoft Information Protection (MIP) sensitivity label.

    Represents the classification-only (unencrypted) label metadata that Office
    stores in ``docMetadata/LabelInfo.xml``. The ``label_id`` and ``site_id``
    identify the label in the tenant's Purview taxonomy; obtain them from the
    Purview portal, from Security & Compliance PowerShell (``Get-Label``), or by
    reading them off a file labeled by hand in Excel via
    :func:`extract_sensitivity_label`.

    Args:
        label_id: The label GUID (with or without surrounding braces).
        site_id: The tenant/site GUID the label belongs to.
        method: ``"Privileged"`` for a label set by explicit user action (what
            Excel writes when a person picks a label), or ``"Standard"`` for an
            automatically applied one. Either satisfies a mandatory-labeling
            policy; defaults to ``"Privileged"`` to match a hand-labeled file.
        content_bits: Visual-marking flags (header/footer/watermark). ``0`` for
            no markings, which is the classification-only case.
        enabled: Whether the label is active. Almost always ``True``.
        removed: Whether the label has been removed. Almost always ``False``.
        name: Optional human-readable label name (e.g. ``"INTERNAL"``).
            Informational only — the modern label format does not store the
            name in the file, so this is not written to the package.
    """

    label_id: str
    site_id: str
    method: str = "Privileged"
    content_bits: int = 0
    enabled: bool = True
    removed: bool = False
    name: Optional[str] = field(default=None, compare=False)


def _braced(guid: str) -> str:
    """Normalize a GUID to the ``{...}`` form Office writes."""
    guid = guid.strip()
    if not guid.startswith("{"):
        guid = "{" + guid
    if not guid.endswith("}"):
        guid = guid + "}"
    return guid


def _labelinfo_xml(label: SensitivityLabel) -> bytes:
    """Render a :class:`SensitivityLabel` as ``LabelInfo.xml`` bytes.

    Built as a literal string rather than via ElementTree so the attribute
    order, self-closing tag, and namespace prefix match byte-for-byte what
    Excel produces.
    """
    return (
        '<?xml version="1.0" encoding="utf-8" standalone="yes"?>'
        f'<clbl:labelList xmlns:clbl="{_LABEL_NS}">'
        f'<clbl:label id="{_braced(label.label_id)}" '
        f'enabled="{1 if label.enabled else 0}" '
        f'method="{label.method}" '
        f'siteId="{_braced(label.site_id)}" '
        f'contentBits="{label.content_bits}" '
        f'removed="{1 if label.removed else 0}" />'
        "</clbl:labelList>"
    ).encode("utf-8")


def extract_sensitivity_label(data: bytes) -> Optional[SensitivityLabel]:
    """Read the sensitivity label from a workbook, if it carries one.

    Args:
        data: Raw .xlsx/.xlsm workbook bytes.

    Returns:
        The :class:`SensitivityLabel` stored in ``docMetadata/LabelInfo.xml``,
        or ``None`` if the file is unlabeled (or not a valid OOXML package).
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            if _LABEL_PART not in zf.namelist():
                return None
            xml = zf.read(_LABEL_PART)
    except zipfile.BadZipFile:
        return None

    try:
        root = ET.fromstring(xml)
    except ET.ParseError:
        return None

    el = root.find(f"{{{_LABEL_NS}}}label")
    if el is None:
        return None

    label_id = el.get("id")
    site_id = el.get("siteId")
    if not label_id or not site_id:
        return None

    try:
        content_bits = int(el.get("contentBits", "0"))
    except ValueError:
        content_bits = 0

    return SensitivityLabel(
        label_id=label_id,
        site_id=site_id,
        method=el.get("method", "Privileged"),
        content_bits=content_bits,
        enabled=el.get("enabled", "1") != "0",
        removed=el.get("removed", "0") != "0",
    )


def _serialize_package_xml(root: ET.Element) -> bytes:
    """Serialize a package XML part with a standard declaration."""
    body = ET.tostring(root, encoding="unicode")
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + body
    ).encode("utf-8")


def _ensure_content_type_override(xml_bytes: bytes) -> bytes:
    """Ensure ``[Content_Types].xml`` declares the label part's content type."""
    ET.register_namespace("", _CT_NS)
    root = ET.fromstring(xml_bytes)
    part_name = "/" + _LABEL_PART
    for override in root.findall(f"{{{_CT_NS}}}Override"):
        if override.get("PartName") == part_name:
            override.set("ContentType", _LABEL_CONTENT_TYPE)
            return _serialize_package_xml(root)
    override = ET.SubElement(root, f"{{{_CT_NS}}}Override")
    override.set("PartName", part_name)
    override.set("ContentType", _LABEL_CONTENT_TYPE)
    return _serialize_package_xml(root)


def _ensure_root_relationship(xml_bytes: bytes) -> bytes:
    """Ensure ``_rels/.rels`` relates the package root to the label part."""
    ET.register_namespace("", _REL_NS)
    root = ET.fromstring(xml_bytes)
    rel_tag = f"{{{_REL_NS}}}Relationship"
    max_id = 0
    for rel in root.findall(rel_tag):
        if rel.get("Type") == _LABEL_REL_TYPE:
            rel.set("Target", _LABEL_PART)
            return _serialize_package_xml(root)
        match = re.fullmatch(r"rId(\d+)", rel.get("Id") or "")
        if match:
            max_id = max(max_id, int(match.group(1)))
    rel = ET.SubElement(root, rel_tag)
    rel.set("Id", f"rId{max_id + 1}")
    rel.set("Type", _LABEL_REL_TYPE)
    rel.set("Target", _LABEL_PART)
    return _serialize_package_xml(root)


def apply_sensitivity_label(data: bytes, label: SensitivityLabel) -> bytes:
    """Stamp a classification-only sensitivity label onto workbook bytes.

    Writes ``docMetadata/LabelInfo.xml`` into the package and wires it into
    ``[Content_Types].xml`` and the root relationships, leaving every other
    part (including ``xl/vbaProject.bin``) untouched. Idempotent: re-applying
    replaces an existing label rather than duplicating it.

    This handles only unencrypted classification labels. It does not apply
    rights-management protection (encryption), which requires the Microsoft
    MIP SDK.

    Args:
        data: Raw .xlsx/.xlsm workbook bytes.
        label: The label to write.

    Returns:
        New workbook bytes carrying the label.
    """
    with zipfile.ZipFile(io.BytesIO(data)) as zin:
        names = zin.namelist()
        entries = {name: zin.read(name) for name in names}
        infos = {name: zin.getinfo(name) for name in names}

    entries[_LABEL_PART] = _labelinfo_xml(label)
    entries["[Content_Types].xml"] = _ensure_content_type_override(
        entries["[Content_Types].xml"]
    )
    entries["_rels/.rels"] = _ensure_root_relationship(entries["_rels/.rels"])

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in names:
            zout.writestr(infos[name], entries[name])
        if _LABEL_PART not in infos:
            zout.writestr(_LABEL_PART, entries[_LABEL_PART])
    return out.getvalue()


def _is_macro_enabled(data: bytes) -> bool:
    """Return True if the workbook bytes are a macro-enabled (.xlsm) file.

    openpyxl only writes the macro-enabled content type when a workbook is
    loaded with ``keep_vba=True``; otherwise it saves the regular .xlsx
    content type. Writing that into a file named ``.xlsm`` makes Excel reject
    it as corrupt ("file format or file extension is not valid").

    Detection is by the package content types, not by the presence of
    ``xl/vbaProject.bin``: a blank macro-enabled template carries the
    macro-enabled content type but has no VBA binary until a macro is added.
    The ``vbaProject.bin`` check is kept as a fallback.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
            if "xl/vbaProject.bin" in names:
                return True
            if "[Content_Types].xml" in names:
                content_types = zf.read("[Content_Types].xml").decode(
                    "utf-8", "ignore"
                )
                return "macroEnabled" in content_types
    except zipfile.BadZipFile:
        return False
    return False


def dataframe_from_excel_bytes(
    data: bytes,
    sheet_name: Optional[Union[str, int]] = None,
) -> pd.DataFrame:
    """Read Excel bytes into a pandas DataFrame.

    Args:
        data: Raw .xlsx file bytes.
        sheet_name: Sheet to read. Defaults to first sheet.

    Returns:
        DataFrame with the sheet data.
    """
    return pd.read_excel(
        io.BytesIO(data),
        engine="openpyxl",
        sheet_name=sheet_name if sheet_name is not None else 0,
    )


def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    sensitivity_label: Optional[SensitivityLabel] = None,
) -> bytes:
    """Write a DataFrame to .xlsx bytes.

    Args:
        df: The DataFrame to write.
        sheet_name: Name of the sheet in the output workbook.
        sensitivity_label: Optional MIP sensitivity label to stamp on the
            output. A DataFrame carries no label of its own, so this is the
            only way to label a workbook written from scratch.

    Returns:
        Raw .xlsx file bytes.
    """
    buf = io.BytesIO()
    df.to_excel(buf, engine="openpyxl", sheet_name=sheet_name, index=False)
    out = buf.getvalue()
    if sensitivity_label is not None:
        out = apply_sensitivity_label(out, sensitivity_label)
    return out


def dataframe_to_excel_bytes_from_template(
    template_data: bytes,
    df: pd.DataFrame,
    sheet_name: Optional[str] = None,
    start_cell: str = "A1",
    orientation: str = "rows",
    include_header: bool = True,
    sensitivity_label: Optional[SensitivityLabel] = None,
) -> bytes:
    """Fill a macro-enabled template with DataFrame data, preserving VBA.

    Loads a ``.xlsm`` template (keeping its VBA project intact), writes the
    DataFrame starting at ``start_cell``, and returns the serialized
    macro-enabled bytes. Use this instead of :func:`dataframe_to_excel_bytes`
    when the template contains macros — re-saving a ``.xlsm`` without its
    VBA project produces a file Excel reports as corrupt.

    Any sensitivity label on the template is preserved on the output, since
    openpyxl would otherwise drop it. Pass ``sensitivity_label`` to set or
    override the label instead.

    Args:
        template_data: Raw .xlsm (or .xlsx) template bytes.
        df: DataFrame to write.
        sheet_name: Target sheet. Defaults to the workbook's active sheet.
            If the named sheet does not exist, a blank single-sheet template is
            renamed to that sheet; otherwise a new sheet is created.
        start_cell: Top-left cell to begin writing (e.g., "A2"). Defaults "A1".
        orientation: "rows" (default) writes each df row as an Excel row;
            "columns" transposes.
        include_header: Whether to write DataFrame column names before the
            values. Defaults True to match ``DataFrame.to_excel(index=False)``.
        sensitivity_label: Optional MIP sensitivity label. When provided, it
            overrides any label already on the template. When omitted, the
            template's own label (if any) is preserved.

    Returns:
        Raw workbook bytes with macros preserved when present in the template.
    """
    template = Template(template_data)
    if sensitivity_label is not None:
        template.set_sensitivity_label(sensitivity_label)
    target_sheet = sheet_name if sheet_name is not None else template.active_sheet
    template.ensure_sheet(target_sheet)
    template.fill_range(
        sheet=target_sheet,
        start_cell=start_cell,
        data=df,
        orientation=orientation,
        include_header=include_header,
    )
    return template.to_bytes()


def _cell_to_row_col(cell_ref: str) -> tuple:
    """Convert a cell reference like 'B3' to (row, col) 1-indexed tuple."""
    try:
        col_letter, row = coordinate_from_string(cell_ref)
    except (ValueError, TypeError) as exc:
        raise ValueError(f"Invalid cell reference: '{cell_ref}'") from exc
    return row, column_index_from_string(col_letter)


class Template:
    """An Excel template that can be populated with data and saved.

    Macro-enabled templates (``.xlsm``) are detected automatically and loaded
    with their VBA project preserved, so macros survive the round-trip and the
    saved file opens cleanly in Excel.

    Args:
        data: Raw .xlsx or .xlsm file bytes of the template.
        keep_vba: Whether to preserve the VBA project. ``None`` (default)
            auto-detects from the file contents. Pass ``True``/``False`` to
            force the behavior.
    """

    def __init__(self, data: bytes, keep_vba: Optional[bool] = None):
        if keep_vba is None:
            keep_vba = _is_macro_enabled(data)
        self._keep_vba = keep_vba
        self._workbook = openpyxl.load_workbook(io.BytesIO(data), keep_vba=keep_vba)
        self._label = extract_sensitivity_label(data)

    @property
    def active_sheet(self) -> str:
        """Name of the workbook's active sheet."""
        return self._workbook.active.title

    @property
    def sensitivity_label(self) -> Optional[SensitivityLabel]:
        """The MIP sensitivity label that will be written on save, if any.

        Initialized from the label the template was loaded with, so a labeled
        template keeps its label through the openpyxl round-trip.
        """
        return self._label

    def set_sensitivity_label(self, label: SensitivityLabel) -> None:
        """Set or override the sensitivity label written on save."""
        self._label = label

    def clear_sensitivity_label(self) -> None:
        """Drop the sensitivity label so the saved file is unlabeled."""
        self._label = None

    def ensure_sheet(self, sheet: str) -> None:
        """Ensure a target sheet exists for a template fill operation.

        A reusable blank template commonly starts with a single empty ``Sheet1``.
        In that case, a requested output sheet name should rename the blank
        sheet instead of leaving an extra empty sheet behind.
        """
        if sheet in self._workbook.sheetnames:
            return

        if (
            len(self._workbook.sheetnames) == 1
            and self._worksheet_is_blank(self._workbook.active)
        ):
            self._workbook.active.title = sheet
            return

        self._workbook.create_sheet(title=sheet)

    @staticmethod
    def _worksheet_is_blank(ws) -> bool:
        return all(cell.value is None for row in ws.iter_rows() for cell in row)

    def fill_range(
        self,
        sheet: Optional[str] = None,
        start_cell: Optional[str] = None,
        end_cell: Optional[str] = None,
        named_range: Optional[str] = None,
        data: Optional[pd.DataFrame] = None,
        orientation: str = "rows",
        allow_expand: bool = False,
        include_header: bool = False,
    ) -> None:
        """Fill a range in the template with DataFrame data.

        Specify either (sheet + start_cell) or named_range, not both.

        Args:
            sheet: Sheet name (required when using start_cell).
            start_cell: Top-left cell to begin writing (e.g., "B3").
            end_cell: Optional bottom-right boundary. Raises if data exceeds it.
            named_range: Name of a defined range in the workbook.
            data: DataFrame to write.
            orientation: "rows" (default) writes each df row as an Excel row.
                "columns" transposes — each df row becomes an Excel column.
            allow_expand: If True, allow writing beyond a named range's bounds.
            include_header: If True, write DataFrame column names before values.
        """
        if data is None:
            raise ValueError("data is required")
        if orientation not in _VALID_ORIENTATIONS:
            raise ValueError(
                f"orientation must be one of {_VALID_ORIENTATIONS}, got '{orientation}'"
            )

        if named_range is not None:
            self._fill_named_range(
                named_range, data, orientation, allow_expand, include_header
            )
        elif sheet is not None and start_cell is not None:
            self._fill_cell_range(
                sheet, start_cell, end_cell, data, orientation, include_header
            )
        else:
            raise ValueError("Provide either (sheet + start_cell) or named_range")

    def _fill_cell_range(
        self,
        sheet: str,
        start_cell: str,
        end_cell: Optional[str],
        data: pd.DataFrame,
        orientation: str,
        include_header: bool,
    ) -> None:
        ws = self._workbook[sheet]
        start_row, start_col = _cell_to_row_col(start_cell)
        end_coords = _cell_to_row_col(end_cell) if end_cell is not None else None
        self._write_block(
            ws,
            start_row,
            start_col,
            data,
            orientation,
            end_coords=end_coords,
            range_label=f"{start_cell}:{end_cell}" if end_cell else None,
            allow_expand=False,
            include_header=include_header,
        )

    def _fill_named_range(
        self,
        range_name: str,
        data: pd.DataFrame,
        orientation: str,
        allow_expand: bool,
        include_header: bool,
    ) -> None:
        defn = self._workbook.defined_names.get(range_name)
        if defn is None:
            raise ValueError(f"Named range '{range_name}' not found in workbook")

        dest_sheet, coord_range = next(iter(defn.destinations))
        ws = self._workbook[dest_sheet]
        parts = coord_range.replace("$", "").split(":")
        start_ref = parts[0]
        end_ref = parts[1] if len(parts) > 1 else None

        start_row, start_col = _cell_to_row_col(start_ref)
        end_coords = _cell_to_row_col(end_ref) if end_ref is not None else None
        self._write_block(
            ws,
            start_row,
            start_col,
            data,
            orientation,
            end_coords=end_coords,
            range_label=f"named range '{range_name}'",
            allow_expand=allow_expand,
            include_header=include_header,
        )

    @staticmethod
    def _write_block(
        ws,
        start_row: int,
        start_col: int,
        data: pd.DataFrame,
        orientation: str,
        end_coords: Optional[tuple] = None,
        range_label: Optional[str] = None,
        allow_expand: bool = False,
        include_header: bool = False,
    ) -> None:
        values = Template._dataframe_to_values(data, orientation, include_header)
        num_rows = len(values)
        num_cols = max((len(row) for row in values), default=0)

        if end_coords is not None and not allow_expand:
            end_row, end_col = end_coords
            max_rows = end_row - start_row + 1
            max_cols = end_col - start_col + 1
            if num_rows > max_rows or num_cols > max_cols:
                msg = (
                    f"Data ({num_rows} rows x {num_cols} cols) exceeds "
                    f"{range_label or 'range'} ({max_rows} rows x {max_cols} cols)"
                )
                if range_label and range_label.startswith("named range"):
                    msg += ". Set allow_expand=True to write beyond the range."
                raise ValueError(msg)

        for r in range(num_rows):
            for c in range(num_cols):
                if c >= len(values[r]):
                    continue
                ws.cell(
                    row=start_row + r,
                    column=start_col + c,
                    value=values[r][c],
                )

    @staticmethod
    def _dataframe_to_values(
        data: pd.DataFrame, orientation: str, include_header: bool
    ) -> list:
        rows = [list(row) for row in data.itertuples(index=False, name=None)]
        if orientation == "rows":
            if include_header:
                return [list(data.columns)] + rows
            return rows

        if include_header:
            return [[col, *data[col].tolist()] for col in data.columns]
        return [list(row) for row in data.to_numpy().T]

    def set_value(self, sheet: str, cell: str, value: object) -> None:
        """Set a single cell value in the template.

        Args:
            sheet: Sheet name.
            cell: Cell reference (e.g., "A1").
            value: Value to write.
        """
        ws = self._workbook[sheet]
        ws[cell] = value

    def to_bytes(self) -> bytes:
        """Serialize the modified template to workbook bytes.

        For macro-enabled templates loaded with ``keep_vba=True``, the VBA
        project is preserved and the output is valid ``.xlsm``. Any sensitivity
        label carried by the template (or set via
        :meth:`set_sensitivity_label`) is re-applied, since openpyxl drops the
        label part on save.
        """
        buf = io.BytesIO()
        self._workbook.save(buf)
        out = buf.getvalue()
        if self._label is not None:
            out = apply_sensitivity_label(out, self._label)
        return out
